#!/usr/bin/env python3
"""
run.py — entry point for social-brain.

Usage:
    python run.py                          # collect + build prompt (default: 2 weeks)
    python run.py --months 3               # collect 3 months of history
    python run.py --collect-only           # collect and save raw data only
    python run.py --analyse-only           # build prompt from most recent saved data
    python run.py --platform <name>        # collect only one platform
    python run.py --update                 # collect + build a compact update prompt for the same chat
    python run.py --analyse-only --update  # update prompt from most recent saved data
    python run.py --extract 2026-03-01     # extract posts from date to today (CSV to stdout)
    python run.py --extract 2026-03-01:2026-03-31  # extract posts in date range
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import sys
from datetime import date, datetime, timedelta, timezone
from html.parser import HTMLParser
from pathlib import Path

import yaml

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

ROOT = Path(__file__).parent
CONFIG_PATH = ROOT / "config.yaml"
DATA_DIR = ROOT / "data" / "weekly"
PLATFORM_DIR = ROOT / "data" / "platform"
REPORTS_DIR = ROOT / "reports"

REQUIRED_CONFIG_KEYS = [
    "mastodon_instance",
    "mastodon_handle",
    "bluesky_handle",
    "buttondown_api_key",
    "jetpack_site",
    "jetpack_access_token",
]


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def load_config() -> dict:
    if not CONFIG_PATH.exists():
        logger.error("config.yaml not found at %s", CONFIG_PATH)
        logger.error(
            "Create it from the template:  cp config.example.yaml config.yaml"
        )
        logger.error("Then fill in your API keys before running social-brain.")
        sys.exit(1)

    with CONFIG_PATH.open() as f:
        config = yaml.safe_load(f)

    if not isinstance(config, dict):
        logger.error("config.yaml is empty or not valid YAML.")
        sys.exit(1)

    missing = [k for k in REQUIRED_CONFIG_KEYS if not config.get(k)]
    if missing:
        logger.warning(
            "The following config keys are missing or empty: %s",
            ", ".join(missing),
        )
        logger.warning(
            "Collectors that need these keys will be skipped or may fail."
        )

    return config


# ---------------------------------------------------------------------------
# Period label  →  e.g.  2025-W22
# ---------------------------------------------------------------------------

def week_label(dt: datetime | None = None, months: int | None = None) -> str:
    if dt is None:
        dt = datetime.now(timezone.utc)
    year, week, _ = dt.isocalendar()
    label = f"{year}-W{week:02d}"
    if months:
        label += f"-{months}m"
    return label


# ---------------------------------------------------------------------------
# File I/O helpers
# ---------------------------------------------------------------------------

def save_raw(data: dict, label: str) -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    path = DATA_DIR / f"{label}.json"
    with path.open("w") as f:
        json.dump(data, f, indent=2, default=str)
    logger.info("Raw data saved → %s", path)
    return path


def save_platform_latest(collected: dict) -> None:
    """Write per-platform latest snapshots to data/platform/<name>-latest.json."""
    PLATFORM_DIR.mkdir(parents=True, exist_ok=True)
    for name, data in collected.items():
        path = PLATFORM_DIR / f"{name}-latest.json"
        with path.open("w") as f:
            json.dump({name: data}, f, indent=2, default=str)
    logger.info("Platform snapshots updated → %s", PLATFORM_DIR)


def load_latest_raw() -> tuple[dict, str]:
    """Return (data, label) for the most recently modified JSON snapshot."""
    snapshots = sorted(DATA_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not snapshots:
        logger.error("No raw data snapshots found in %s", DATA_DIR)
        logger.error("Run without --analyse-only first to collect data.")
        sys.exit(1)
    path = snapshots[0]
    label = path.stem
    logger.info("Loading raw data from %s", path)
    with path.open() as f:
        return json.load(f), label


def check_drop_staleness(config: dict | None = None) -> list[str]:
    """
    Check file-drop directories for stale exports.
    Returns a list of warning strings for any that are out of date.
    Only warns if files exist (i.e. the user has previously dropped exports).
    Pass config to suppress LinkedIn staleness warning when API token is configured.
    """
    now = datetime.now(timezone.utc)
    warnings: list[str] = []

    def _most_recent(*globs: str) -> Path | None:
        files = []
        for g in globs:
            parts = g.rsplit("/", 1)
            if len(parts) == 2:
                files.extend(Path(parts[0]).glob(parts[1]) if Path(parts[0]).exists() else [])
            else:
                files.extend(Path(".").glob(g))
        return max(files, key=lambda p: p.stat().st_mtime) if files else None

    # LinkedIn: must be < 24 hours old (skip if API token is configured)
    if not (config and config.get("linkedin_access_token")):
        newest = _most_recent("linkedin_drops/*.csv", "linkedin_drops/*.xlsx")
        if newest:
            age = now - datetime.fromtimestamp(newest.stat().st_mtime, tz=timezone.utc)
            if age > timedelta(hours=24):
                hours = age.total_seconds() / 3600
                warnings.append(
                    f"LinkedIn: export '{newest.name}' is {hours:.0f}h old — download a fresh export first."
                )

    # O'Reilly: warn if most recent statement is > 25 days old (monthly payment cycle)
    newest_or = _most_recent("oreilly_drops/*.eml", "oreilly_drops/*.rtf")
    if newest_or:
        age = now - datetime.fromtimestamp(newest_or.stat().st_mtime, tz=timezone.utc)
        if age > timedelta(days=25):
            warnings.append(
                f"O'Reilly: most recent statement '{newest_or.name}' is {age.days} days old — "
                f"check your email for a new remittance statement."
            )

    return warnings


def _platform_expected(name: str, config: dict) -> bool:
    """
    Return True if a platform was expected to collect data — meaning it was
    configured (API key/token present) or its file-drop directory has files.
    Platforms that are neither configured nor have drop files are silently ignored.
    """
    def _has_files(*globs: str) -> bool:
        for g in globs:
            parts = g.rsplit("/", 1)
            if len(parts) == 2:
                base = Path(parts[0])
                if base.exists() and any(base.glob(parts[1])):
                    return True
            elif any(Path(".").glob(g)):
                return True
        return False

    if name == "mastodon":
        return bool(config.get("mastodon_instance") and config.get("mastodon_handle"))
    if name == "bluesky":
        return bool(config.get("bluesky_handle"))
    if name == "buttondown":
        return bool(config.get("buttondown_api_key"))
    if name == "jetpack":
        return bool(config.get("jetpack_site") and config.get("jetpack_access_token"))
    if name == "linkedin":
        return bool(config.get("linkedin_access_token")) or _has_files("linkedin_drops/*.csv", "linkedin_drops/*.xlsx")
    if name == "substack":
        return _has_files("substack_drops/*.csv")
    if name == "vercel":
        return bool(config.get("vercel_token") and config.get("vercel_project_id"))
    if name == "amazon":
        return bool(config.get("amazon_asins"))
    if name == "upcoming":
        return bool(
            (config.get("jetpack_site") and config.get("jetpack_access_token"))
            or config.get("buttondown_api_key")
            or config.get("buffer_token")
        )
    if name == "mentions":
        return bool(config.get("monitored_domains"))
    if name == "goatcounter":
        return bool(config.get("goatcounter_site") and config.get("goatcounter_token"))
    if name == "oreilly":
        return _has_files("oreilly_drops/*.eml", "oreilly_drops/*.rtf")
    if name == "calendly":
        return bool(config.get("calendly_token"))
    return False


def _platform_summary(name: str, data: dict) -> str:
    """Return a brief one-line description of what was collected for a platform."""
    try:
        if name == "mastodon":
            followers = data.get("account", {}).get("followers", "?")
            posts = len(data.get("posts", []))
            return f"{followers} followers, {posts} posts"
        if name == "bluesky":
            return f"{len(data.get('posts', []))} posts"
        if name == "buttondown":
            return f"{len(data.get('newsletters', []))} newsletter(s)"
        if name == "jetpack":
            return f"{data.get('total_views', '?')} views"
        if name == "linkedin":
            days = len(data.get("daily_engagement", []))
            posts = len(data.get("top_posts_by_engagement", []))
            return f"{days} days engagement, {posts} top posts"
        if name == "substack":
            return f"{len(data.get('emails', []))} emails"
        if name == "vercel":
            return f"{data.get('page_views', '?')} views"
        if name == "amazon":
            return f"{len(data.get('by_marketplace', {}))} marketplace(s)"
        if name == "goatcounter":
            return f"{data.get('total_pageviews', '?')} pageviews"
        if name == "oreilly":
            count = data.get("payment_count", 0)
            total = data.get("total_paid", 0.0)
            currencies = data.get("currencies") or ["?"]
            return f"{count} payment(s), {currencies[0]} {total:.2f} total"
        if name == "calendly":
            bookings = data.get("total_bookings", 0)
            lead = data.get("lead_gen_bookings")
            if lead is not None:
                return f"{bookings} booking(s) ({lead} lead gen)"
            return f"{bookings} booking(s)"
        if name == "mentions":
            sources = data.get("sources", {})
            hn = len(sources.get("hackernews", []))
            mastodon = len(sources.get("mastodon", []))
            bluesky = len(sources.get("bluesky", []))
            return f"HN: {hn}, Mastodon: {mastodon}, Bluesky: {bluesky}"
        if name == "upcoming":
            sources = data.get("sources", {})
            total = sum(len(v) for v in sources.values() if isinstance(v, list))
            return f"{total} scheduled item(s)"
    except Exception:
        pass
    return "ok"


def _print_run_summary(
    collected: dict,
    all_platforms: list[str],
    config: dict,
    prompt_path: Path | None = None,
    update: bool = False,
) -> None:
    """Print a clean end-of-run summary: what was collected and what wasn't."""
    print("\n" + "=" * 52)
    print("  Run complete")
    print("=" * 52)

    present = [p for p in all_platforms if p in collected]
    # Only flag as "not collected" platforms that were actually expected to have data
    missing = [
        p for p in all_platforms
        if p not in collected and _platform_expected(p, config)
    ]

    if present:
        print("\nCollected:")
        for p in present:
            print(f"  ✓  {p:<14} {_platform_summary(p, collected[p])}")

    if missing:
        print("\nExpected but not collected:")
        for p in missing:
            print(f"  –  {p}")

    if prompt_path:
        action = "Paste into your existing claude.ai chat" if update else "Paste into claude.ai"
        size_kb = f"{prompt_path.stat().st_size // 1024} KB" if prompt_path.exists() else ""
        size_str = f" ({size_kb})" if size_kb else ""
        print(f"\nPrompt{size_str}: {prompt_path}")
        print(f"→  {action}")

    print("=" * 52 + "\n")


def extract_posts(date_range: str) -> None:
    """Extract posts from analytics.xlsx and write CSV to stdout."""

    parts = date_range.split(":")
    try:
        start = datetime.strptime(parts[0].strip(), "%Y-%m-%d").date()
        end = datetime.strptime(parts[1].strip(), "%Y-%m-%d").date() if len(parts) > 1 else datetime.now(timezone.utc).date()
    except ValueError:
        logger.error("Invalid date range '%s'. Expected YYYY-MM-DD or YYYY-MM-DD:YYYY-MM-DD", date_range)
        sys.exit(1)

    store_path = ROOT / "data" / "analytics.xlsx"
    if not store_path.exists():
        logger.error("analytics.xlsx not found — run without --extract first to collect data.")
        sys.exit(1)

    import openpyxl

    class _HTMLStripper(HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self._parts: list[str] = []
        def handle_data(self, data: str) -> None:
            self._parts.append(data)
        def stripped(self) -> str:
            return " ".join(self._parts).strip()

    def strip_html(html: str) -> str:
        s = _HTMLStripper()
        s.feed(html)
        return s.stripped()

    def parse_date(val: object) -> date | None:
        if val is None:
            return None
        s = str(val).strip()
        for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(s[: len(fmt) + 4], fmt).date()
            except ValueError:
                continue
        return None

    wb = openpyxl.load_workbook(store_path)
    posts: list[dict] = []

    def _read_sheet(sheet_name: str, date_col: str, text_col: str, platform: str, html: bool = False) -> None:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            r = dict(zip(headers, row))
            d = parse_date(r.get(date_col))
            if d and start <= d <= end:
                text = r.get(text_col) or ""
                posts.append({
                    "date": d,
                    "platform": platform,
                    "text": strip_html(text) if html else str(text).strip(),
                })

    _DATE_COLS = ["created_at", "date", "send_date", "indexed_at"]
    _TEXT_COLS = ["content", "text", "subject"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header_set = {h for h in headers if h}
        date_col = next((c for c in _DATE_COLS if c in header_set), None)
        text_col = next((c for c in _TEXT_COLS if c in header_set), None)
        if date_col and text_col:
            _read_sheet(sheet_name, date_col, text_col, sheet_name, html=(text_col == "content"))

    posts.sort(key=lambda p: p["date"])

    writer = csv.writer(sys.stdout)
    writer.writerow(["date", "platform", "text"])
    for p in posts:
        writer.writerow([p["date"], p["platform"], p["text"]])

    logger.info("Extracted %d post(s) from %s to %s", len(posts), start, end)


def since_last_run() -> datetime | None:
    """
    Return the mtime of the most recent snapshot if it's older than 2 weeks,
    so the caller can extend the lookback window to cover the gap.
    Returns None if the last run was within the default window.
    """
    snapshots = sorted(DATA_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not snapshots:
        return None
    last_mtime = datetime.fromtimestamp(snapshots[0].stat().st_mtime, tz=timezone.utc)
    if last_mtime < datetime.now(timezone.utc) - timedelta(weeks=2):
        return last_mtime
    return None



# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# OAuth helpers
# ---------------------------------------------------------------------------

def _linkedin_oauth(config: dict) -> None:
    """
    Run the LinkedIn 3-legged OAuth flow:
      1. Print the authorization URL and open it in the browser.
      2. Start a one-shot localhost HTTP server to receive the callback.
      3. Exchange the code for an access token.
      4. Write the token to config.yaml.

    Requires config keys: linkedin_client_id, linkedin_client_secret.
    Token is written to: linkedin_access_token.
    """
    import http.server
    import threading
    import secrets
    import urllib.parse
    import webbrowser

    client_id = config.get("linkedin_client_id", "")
    client_secret = config.get("linkedin_client_secret", "")
    if not client_id or not client_secret:
        logger.error(
            "LinkedIn OAuth requires linkedin_client_id and linkedin_client_secret in config.yaml.\n"
            "See README for setup instructions."
        )
        sys.exit(1)

    redirect_uri = "http://localhost:8976/callback"
    scope = "r_dma_portability_self_serve"
    state = secrets.token_urlsafe(16)

    auth_url = (
        "https://www.linkedin.com/oauth/v2/authorization"
        f"?response_type=code"
        f"&client_id={urllib.parse.quote(client_id)}"
        f"&redirect_uri={urllib.parse.quote(redirect_uri)}"
        f"&scope={urllib.parse.quote(scope)}"
        f"&state={urllib.parse.quote(state)}"
    )

    # Capture the callback code via a one-shot HTTP server
    received: dict = {}
    server_ready = threading.Event()

    class _Handler(http.server.BaseHTTPRequestHandler):
        def do_GET(self) -> None:  # noqa: N802
            parsed = urllib.parse.urlparse(self.path)
            params = dict(urllib.parse.parse_qsl(parsed.query))
            received.update(params)
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.end_headers()
            self.wfile.write(
                b"<html><body><h2>social-brain: LinkedIn authorised!</h2>"
                b"<p>You can close this tab and return to the terminal.</p></body></html>"
            )

        def log_message(self, *args: object) -> None:  # suppress server logs
            pass

    server = http.server.HTTPServer(("localhost", 8976), _Handler)
    thread = threading.Thread(target=server.handle_request, daemon=True)
    thread.start()

    print(f"\nOpening LinkedIn OAuth page in your browser...")
    print(f"If it doesn't open automatically, visit:\n  {auth_url}\n")
    webbrowser.open(auth_url)

    print("Waiting for LinkedIn to redirect to localhost:8976 ...")
    thread.join(timeout=120)

    if not received.get("code"):
        logger.error("OAuth timed out or no code received. Try again.")
        sys.exit(1)

    if received.get("state") != state:
        logger.error("OAuth state mismatch — possible CSRF. Aborting.")
        sys.exit(1)

    code = received["code"]

    # Exchange code for token
    import httpx as _httpx
    try:
        r = _httpx.post(
            "https://www.linkedin.com/oauth/v2/accessToken",
            data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": redirect_uri,
                "client_id": client_id,
                "client_secret": client_secret,
            },
            timeout=30,
        )
        r.raise_for_status()
    except Exception as exc:
        logger.error("Token exchange failed: %s", exc)
        sys.exit(1)

    token_data = r.json()
    access_token = token_data.get("access_token")
    if not access_token:
        logger.error("No access_token in response: %s", token_data)
        sys.exit(1)

    # Write token to config.yaml
    with CONFIG_PATH.open() as f:
        raw_yaml = f.read()

    if "linkedin_access_token:" in raw_yaml:
        import re as _re
        raw_yaml = _re.sub(
            r"^linkedin_access_token:.*$",
            f"linkedin_access_token: {access_token}",
            raw_yaml,
            flags=_re.MULTILINE,
        )
    else:
        raw_yaml += f"\nlinkedin_access_token: {access_token}\n"

    with CONFIG_PATH.open("w") as f:
        f.write(raw_yaml)

    expires_in = token_data.get("expires_in", 5183999)
    expires_days = expires_in // 86400
    print(f"\nLinkedIn token saved to config.yaml (valid for ~{expires_days} days).")
    print("Run `python run.py --platform linkedin` to verify collection works.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="social-brain: collect social analytics and generate a weekly report."
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--collect-only",
        action="store_true",
        help="Collect data and save raw JSON, but skip analysis.",
    )
    mode.add_argument(
        "--analyse-only",
        action="store_true",
        help="Skip collection and analyse the most recent saved raw data.",
    )
    mode.add_argument(
        "--extract",
        metavar="RANGE",
        help="Extract posts from analytics.xlsx and write CSV to stdout. "
             "RANGE is YYYY-MM-DD (from date to today) or YYYY-MM-DD:YYYY-MM-DD.",
    )
    from collectors import PLATFORM_COLLECTORS
    parser.add_argument(
        "--platform",
        choices=sorted(PLATFORM_COLLECTORS.keys()),
        default=None,
        help="Collect only one platform (cannot be combined with --analyse-only).",
    )
    parser.add_argument(
        "--months",
        type=int,
        default=None,
        metavar="N",
        help="Collect N months of history instead of the default 2-week window.",
    )
    parser.add_argument(
        "--update",
        action="store_true",
        help="Generate a compact follow-up prompt (prompt-YYYY-WNN-update.txt) for use in the same claude.ai chat as the original report.",
    )
    mode.add_argument(
        "--auth",
        metavar="PLATFORM",
        choices=["linkedin"],
        help="Authenticate with a platform via OAuth and save the token to config.yaml. "
             "Currently supports: linkedin",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.extract:
        extract_posts(args.extract)
        return

    if args.auth:
        config = load_config()
        if args.auth == "linkedin":
            _linkedin_oauth(config)
        return

    if args.analyse_only and args.platform:
        logger.error("--analyse-only and --platform cannot be used together.")
        sys.exit(1)

    config = load_config()

    since: datetime | None = None
    if args.months:
        since = datetime.now(timezone.utc) - timedelta(days=args.months * 30)
        logger.info("Lookback: %d month(s) (since %s)", args.months, since.date())
    elif not args.analyse_only:
        gap = since_last_run()
        if gap:
            since = gap
            logger.info("Last run was %s — extending lookback to cover gap", gap.date())

    label = week_label(months=args.months)

    # ------------------------------------------------------------------
    # Staleness check
    # ------------------------------------------------------------------
    if not args.analyse_only:
        stale = check_drop_staleness(config)
        if stale:
            print("\nWarning: stale data detected:")
            for w in stale:
                print(f"  • {w}")
            if not sys.stdin.isatty():
                logger.warning("Running non-interactively — continuing despite stale data.")
            else:
                try:
                    answer = input("\nContinue anyway? [y/N] ").strip().lower()
                except (EOFError, KeyboardInterrupt):
                    answer = ""
                if answer != "y":
                    logger.info("Aborted by user.")
                    sys.exit(0)

    # ------------------------------------------------------------------
    # Collection
    # ------------------------------------------------------------------
    collected: dict = {}

    if not args.analyse_only:
        from collect import collect_all

        logger.info("=== Collecting data (label: %s) ===", label)
        collected = collect_all(config, platform=args.platform, since=since)

        if not collected:
            logger.warning("No data was collected from any platform.")
        else:
            logger.info(
                "Collected data from: %s", ", ".join(collected.keys())
            )

        save_raw(collected, label)
        if collected:
            save_platform_latest(collected)

        # ------------------------------------------------------------------
        # Persistent store — upsert into analytics.xlsx
        # ------------------------------------------------------------------
        if collected and not args.platform:
            from store import update as store_update, get_known_platforms, STORE_PATH

            known = get_known_platforms()
            new_platforms = set(collected.keys()) - known - {"upcoming", "mentions"}

            if new_platforms and since is None:
                # First time seeing these platforms — backfill 3 months
                logger.info(
                    "Store: new platform(s) detected (%s) — backfilling 3 months",
                    ", ".join(sorted(new_platforms)),
                )
                backfill_since = datetime.now(timezone.utc) - timedelta(days=90)
                backfill = collect_all(
                    config,
                    platform=None,
                    since=backfill_since,
                )
                store_update(backfill)
            else:
                store_update(collected)

        if args.collect_only:
            from collectors import PLATFORM_COLLECTORS
            _print_run_summary(collected, list(PLATFORM_COLLECTORS.keys()), config)
            return

    # ------------------------------------------------------------------
    # Analysis
    # ------------------------------------------------------------------
    if args.analyse_only:
        collected, label = load_latest_raw()

    if not collected:
        logger.warning(
            "No collected data to analyse — the report will be minimal."
        )

    logger.info("=== Building analysis prompt ===")
    from analyse import save_prompt
    from collectors import PLATFORM_COLLECTORS

    prompt_path = save_prompt(collected, config, period=label, reports_dir=REPORTS_DIR, months=args.months, update=args.update)
    _print_run_summary(collected, list(PLATFORM_COLLECTORS.keys()), config, prompt_path=prompt_path, update=args.update)


if __name__ == "__main__":
    main()
